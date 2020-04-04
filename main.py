import requests
import re
from tqdm import tqdm
from bs4 import BeautifulSoup
import openpyxl
from tqdm import tqdm

wb = openpyxl.load_workbook('data.xlsx')
sheet = wb.active
sheet.title = 'UK Indie Retailers'
EMAIL_REGEX = r'[^ ]+@[^ ]+'

def scrape_data():
    retailer_counter = 2
    counter = 1

    for i in tqdm(range(2, 154)):

        url = f'https://www.indieretail.uk/find-a-shop/?page={counter}'

        page = requests.get(url)
        soup = BeautifulSoup(page.content, 'html.parser')


        data_div = soup.find_all('div', {'class': 'shoplisting'})

        if len(data_div) == 1:
            if data_div[0].text.strip().split()[0].lower() == 'sorry':
                break


        for div in data_div:
            data = div.find('a', {'itemprop':'url'}, href=True)

            company_name = data.text

            href = f'https://www.indieretail.uk/find-a-shop{data["href"][1:]}'

            shop_page = requests.get(href)
            shop_page_soup = BeautifulSoup(shop_page.content, 'html.parser')

            company_website = shop_page_soup.find('a', {'itemprop':'url','target':'_blank'})

            if company_website is None:
                company_website = "-"
            else:
                company_website = company_website.text

            cell_1 = sheet.cell(row=retailer_counter, column=1)
            cell_1.value = company_name
            cell_2 = sheet.cell(row=retailer_counter, column=2)
            cell_2.value = company_website
            cell_3 = sheet.cell(row=retailer_counter, column=3)

            company_email = '-'

            try:
                hdr = {'User-Agent': 'Mozilla/5.0'}
                company_page = requests.get(f'http://{company_website}', headers=hdr, timeout=5)

                if company_page.status_code != 404:
                    company_soup = BeautifulSoup(company_page.content, 'html.parser')
                    company_links = company_soup.findAll('a')
                    for link in company_links:
                        text = link.text
                        if len(re.findall(EMAIL_REGEX, text)) != 0:
                            company_email = text
                            break

                cell_3.value = company_email

            except:
                cell_3.value = company_email

            finally:
                cell_3.value = company_email

            wb.save('complete_data.xlsx')

            retailer_counter += 1

        counter += 1



def main():
    scrape_data()

if __name__ == '__main__':
    main()
